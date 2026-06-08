"""Extract recipe data from Nendys V2 spreadsheet into JSON.

Approach:
- Build a material map: cell coord (e.g. 'I12') -> material id (e.g. 'STEEL_T1')
- For every craft sheet, scan recipe rows for cost formulas that reference Materials!$X$Y.
- Parse each formula into: list of {material_id, qty}, item tier label, enchantment.
- Output single JSON consumed by the website.
"""
import json
import re
from openpyxl import load_workbook

PATH = r'C:\Users\Bernardo\Downloads\Nendys V2 (7).xlsx'
OUT = r'C:\Users\Bernardo\Desktop\AlbionSite\albion\data.json'

wb = load_workbook(PATH, data_only=False)


# ---------- ITEM VALUE (IV) FORMULA ----------
# Standard Albion rule, confirmed against the game's items.json:
# refined materials at T4.0 have IV=16 (T4_METALBAR, T4_LEATHER, T4_CLOTH,
# T4_STONEBLOCK), and IV doubles with each tier and each enchantment level.
#   IV(tier, ench) = 16 * 2^(tier - 4) * 2^ench
# This is exact for refined resources / raw resources / crafted equipment
# (where item IV ≈ sum of input material IVs). Food / potion ingredients
# (crops, herbs, eggs) follow the same scaling in the absence of better data.

def derive_iv(tier_str):
    """Tier string like 'T4.0', 'T8.4', 'T1', 'T3' -> numeric IV."""
    if not tier_str or tier_str == '-':
        return 0
    s = str(tier_str).strip().upper().lstrip('T')
    try:
        if '.' in s:
            t, e = s.split('.', 1)
            tier, ench = int(t), int(e)
        else:
            tier, ench = int(s), 0
    except ValueError:
        return 0
    return 16 * (2 ** (tier - 4)) * (2 ** ench)


def parse_tier_from_name(name):
    """Pull a tier number out of names like 'TIER 3 - Sheaf of Wheat'."""
    if not name:
        return None
    import re as _re
    m = _re.search(r'TIER\s*(\d+)', str(name).upper())
    return f'T{m.group(1)}' if m else None


# ---------- BUILD MATERIAL MAP ----------
# Materials sheet structure:
# Refined (rows 9-66): B/G/L/Q/V are labels; D/I/N/S/X are prices.
#   Row 9 headers: B=PLANKS G=STEEL L=LEATHER Q=CLOTH V=BLOCKS
#   Tier rows: 12=T1, 14=T2, 16=T3, 18=T4.0, 20=T4.1, 22=T4.2, 24=T4.3, 26=T4.4,
#              28=T5.0, 30=T5.1, 32=T5.2, 34=T5.3, 36=T5.4,
#              38=T6.0, 40=T6.1, 42=T6.2, 44=T6.3, 46=T6.4,
#              48=T7.0, 50=T7.1, 52=T7.2, 54=T7.3, 56=T7.4,
#              58=T8.0, 60=T8.1, 62=T8.2, 64=T8.3, 66=T8.4
# Stone shifts: V12-V26 are T1-T8 only (basic), but matches B12-B26 same pattern in "raw" section.
# Raw (rows 78-135) same column layout.
# Special: column X (rows 33,35,37,39,41,43,45) = hearts; V52,V54,V56 = Tome/Siphoned/Avalonian; V58-66 = sigils.

REFINED_TIERS = {
    12: 'T1', 14: 'T2', 16: 'T3',
    18: 'T4.0', 20: 'T4.1', 22: 'T4.2', 24: 'T4.3', 26: 'T4.4',
    28: 'T5.0', 30: 'T5.1', 32: 'T5.2', 34: 'T5.3', 36: 'T5.4',
    38: 'T6.0', 40: 'T6.1', 42: 'T6.2', 44: 'T6.3', 46: 'T6.4',
    48: 'T7.0', 50: 'T7.1', 52: 'T7.2', 54: 'T7.3', 56: 'T7.4',
    58: 'T8.0', 60: 'T8.1', 62: 'T8.2', 64: 'T8.3', 66: 'T8.4',
}
RAW_TIERS = {
    81: 'T1', 83: 'T2', 85: 'T3',
    87: 'T4.0', 89: 'T4.1', 91: 'T4.2', 93: 'T4.3', 95: 'T4.4',
    97: 'T5.0', 99: 'T5.1', 101: 'T5.2', 103: 'T5.3', 105: 'T5.4',
    107: 'T6.0', 109: 'T6.1', 111: 'T6.2', 113: 'T6.3', 115: 'T6.4',
    117: 'T7.0', 119: 'T7.1', 121: 'T7.2', 123: 'T7.3', 125: 'T7.4',
    127: 'T8.0', 129: 'T8.1', 131: 'T8.2', 133: 'T8.3', 135: 'T8.4',
}

# (price_col, family) per refined section
REFINED_COLS = {'D': 'PLANKS', 'I': 'STEEL', 'N': 'LEATHER', 'S': 'CLOTH', 'X': 'BLOCKS'}
RAW_COLS = {'D': 'LOGS', 'I': 'ORE', 'N': 'HIDE', 'S': 'FIBER', 'X': 'STONE'}

mat_map = {}      # 'I12' -> id
mat_meta = {}     # id -> {family,tier,name,kind}

# Stone block (refined) has NO enchant variants in Albion — the V column
# in the refined section therefore labels rows 12-26 as T1..T8 with no
# decimal suffix. The standard REFINED_TIERS map would (wrongly) create
# BLOCKS_T4.0/T4.1/...
BLOCKS_TIERS = {
    12: 'T1', 14: 'T2', 16: 'T3',
    18: 'T4', 20: 'T5', 22: 'T6', 24: 'T7', 26: 'T8',
}
# Raw stone (rocks) has only 4 enchant variants per tier 4+ (base + LEVEL1-3,
# no LEVEL4). The V column in the raw section therefore "shifts" relative
# to logs/ore/hide/fiber.
STONE_TIERS = {
    81: 'T1', 83: 'T2', 85: 'T3',
    87: 'T4.0',  89: 'T4.1',  91: 'T4.2',  93: 'T4.3',
    95: 'T5.0',  97: 'T5.1',  99: 'T5.2', 101: 'T5.3',
    103:'T6.0', 105: 'T6.1', 107: 'T6.2', 109: 'T6.3',
    111:'T7.0', 113: 'T7.1', 115: 'T7.2', 117: 'T7.3',
    119:'T8.0', 121: 'T8.1', 123: 'T8.2', 125: 'T8.3',
    # Stone Refining T8 ench 3 formula in the spreadsheet references X127,
    # which would be a non-existent T8.4 rock. Re-route the typo to T8.3
    # so the recipe resolves to the correct material id.
    127: 'T8.3',
}

# Refined
for col, family in REFINED_COLS.items():
    tiers = BLOCKS_TIERS if family == 'BLOCKS' else REFINED_TIERS
    for row, tier in tiers.items():
        mid = f'{family}_{tier}'
        mat_map[f'{col}{row}'] = mid
        mat_meta[mid] = {'id': mid, 'family': family, 'tier': tier, 'kind': 'refined',
                         'name': f'{family.title()} {tier}'}

# Raw
for col, family in RAW_COLS.items():
    tiers = STONE_TIERS if family == 'STONE' else RAW_TIERS
    for row, tier in tiers.items():
        mid = f'{family}_{tier}'
        mat_map[f'{col}{row}'] = mid
        mat_meta[mid] = {'id': mid, 'family': family, 'tier': tier, 'kind': 'raw',
                         'name': f'{family.title()} {tier}'}

# Hearts (column X, rows 33-45 odd-ish in refined section)
HEARTS = {33: 'Beastheart', 35: 'Mountainheart', 37: 'Treeheart',
          39: 'Rockheart', 41: 'Vineheart', 43: 'Faerie Fire', 45: 'Shadowheart'}
for row, name in HEARTS.items():
    mid = f'HEART_{name.upper().replace(" ","_")}'
    mat_map[f'X{row}'] = mid
    mat_meta[mid] = {'id': mid, 'family': 'HEART', 'tier': '-', 'kind': 'heart', 'name': name}

# Misc (V column rows 52,54,56,58-66)
MISC = {
    52: 'Tome of Insight', 54: 'Siphoned Energy', 56: 'Avalonian Energy',
    58: 'Royal Sigil T4', 60: 'Royal Sigil T5', 62: 'Royal Sigil T6',
    64: 'Royal Sigil T7', 66: 'Royal Sigil T8',
}
for row, name in MISC.items():
    mid = 'MISC_' + name.upper().replace(' ', '_').replace('.', '')
    mat_map[f'X{row}'] = mid
    mat_meta[mid] = {'id': mid, 'family': 'MISC', 'tier': '-', 'kind': 'misc', 'name': name}

# Food & Potion materials (rows 148-210)
# Each label column (B, G, L, Q, V) carries multiple sub-categories stacked
# vertically. UPPERCASE rows are sub-category headers (CROPS, HERBS, MILK,
# EGGS, BREWING, WHEAT PRODUCTS, RAW MEAT, ANIMALS, SHADOW CLAWS, SYLVIAN
# ROOT, WEREWOLF FANGS, SPIRIT PAWS, IMP'S HORNS, RUNESTONE TOOTH, DAWNFEATHER,
# OTHER ITEMS). Items below each header inherit it as their sub-category.
ws = wb['Materials']
for c_letter, p_letter in [('B','D'), ('G','I'), ('L','N'), ('Q','S'), ('V','X')]:
    current_sub = 'OTHER'
    for r in range(140, 210):
        name_cell = ws[f'{c_letter}{r}'].value
        if not (name_cell and isinstance(name_cell, str) and name_cell.strip()):
            continue
        v = name_cell.strip()
        # Sub-category header — uppercase short string
        if v.isupper() and len(v) <= 20:
            current_sub = v
            continue
        mid = 'FP_' + v.upper().replace(' ', '_').replace('-', '_').replace("'", '').replace(',', '')
        mat_map[f'{p_letter}{r}'] = mid
        mat_meta[mid] = {
            'id': mid,
            'family': 'FOOD_POTION',
            'subFamily': current_sub,
            'tier': '-',
            'kind': 'foodpotion',
            'name': v,
        }

# Enchantment materials (V/X starting around 186-202: extracts, fish products)
# Already covered by the loop above.

# Hideout silver (V204)
mat_map['X204'] = 'MISC_HIDEOUT_SILVER'
mat_meta['MISC_HIDEOUT_SILVER'] = {'id': 'MISC_HIDEOUT_SILVER', 'family': 'MISC',
                                    'tier': '-', 'kind': 'misc', 'name': 'Hideout Silver Cost'}


# ---------- PARSE RECIPE FORMULAS ----------
# Standard pattern: qty * Materials!$col$row
MAT_REF = re.compile(r"(\d+(?:\.\d+)?)\s*\*\s*Materials!\$?([A-Z])\$?(\d+)")
# Hearts-conditional pattern: MAX(N - IF($I$21,1,0),0) * Materials!$col$row
HEART_COND_REF = re.compile(
    r"MAX\(\s*(\d+)\s*-\s*IF\(\$?I\$?\d+\s*,\s*1\s*,\s*0\)\s*,\s*0\)\s*\*\s*Materials!\$?([A-Z])\$?(\d+)",
    re.IGNORECASE,
)
# Plain heart reference: IF($I$21,Materials!$X$33,0)
HEART_IF_REF = re.compile(
    r"IF\(\$?I\$?\d+\s*,\s*Materials!\$?([A-Z])\$?(\d+)\s*,\s*0\)",
    re.IGNORECASE,
)
# Any leftover Materials! reference (qty assumed 1)
SIMPLE_REF = re.compile(r"Materials!\$?([A-Z])\$?(\d+)")

def parse_formula(formula):
    """Returns list of {mat, qty, heartReducesQty?, heartGated?, noReturnDiscount?}.

    - heartReducesQty: when useHearts is on, subtract 1 from qty.
    - heartGated: when useHearts is off, omit (qty *= 0).
    - noReturnDiscount: ingredient lives OUTSIDE the (1 - returnFactor) bracket.
      In Albion this means specialty drops (Runestone Tooth, Imp's Horns, etc.)
      that the crafting station's return rate doesn't refund.
    """
    if not formula or not isinstance(formula, str):
        return None
    f = formula.strip()
    if not f.startswith('='):
        return None

    # ---- detect (1 - <ref>) factor & batch-divisor positions ----
    # Anything between the factor's end and the batch-divisor `/N` lives OUTSIDE
    # the discount bracket and must NOT be multiplied by (1 - returnFactor).
    factor_match = re.search(r"\(\s*1\s*-\s*\$?[A-Z]\$?\d+\s*\)", f)
    factor_end   = factor_match.end() if factor_match else None
    batch_pos    = None
    for m in re.finditer(r"/\s*(\d+(?:\.\d+)?)", f):
        n = float(m.group(1))
        if 1 < n <= 50:
            batch_pos = m.start()
            break
    def outside_discount(pos):
        if factor_end is None or pos <= factor_end:
            return False
        return batch_pos is None or pos < batch_pos

    items = []
    found = set()  # coords already captured

    # 1) Heart-conditional quantity: MAX(N - IF($I$21,1,0),0) * Materials!XY
    for m in HEART_COND_REF.finditer(f):
        baseQty = float(m.group(1))
        coord = f'{m.group(2)}{m.group(3)}'
        mid = mat_map.get(coord)
        if mid:
            items.append({'mat': mid, 'qty': baseQty, 'heartReducesQty': True})
            found.add(coord)

    # 2) Standard qty * Materials!XY
    for m in MAT_REF.finditer(f):
        coord = f'{m.group(2)}{m.group(3)}'
        if coord in found:
            continue
        qty = float(m.group(1))
        mid = mat_map.get(coord)
        if mid:
            it = {'mat': mid, 'qty': qty}
            if outside_discount(m.start()):
                it['noReturnDiscount'] = True
            items.append(it)
            found.add(coord)

    # 3) Heart references (gated by useHearts toggle): IF($I$21, Materials!$X$33, 0)
    for m in HEART_IF_REF.finditer(f):
        coord = f'{m.group(1)}{m.group(2)}'
        if coord in found:
            continue
        mid = mat_map.get(coord)
        if mid:
            items.append({'mat': mid, 'qty': 1, 'heartGated': True})
            found.add(coord)

    # 4) Catch-all refs without explicit multiplier (qty=1)
    for m in SIMPLE_REF.finditer(f):
        coord = f'{m.group(1)}{m.group(2)}'
        if coord in found:
            continue
        mid = mat_map.get(coord)
        if mid:
            it = {'mat': mid, 'qty': 1}
            if outside_discount(m.start()):
                it['noReturnDiscount'] = True
            items.append(it)
            found.add(coord)

    return items if items else None


# Station-fee pattern from Food/Potions formulas:
#   + (NUTRITION * 0.1125 * C16 / 100)
# C16 holds the user-input station fee. NUTRITION is recipe-specific.
STATION_FEE_RE = re.compile(
    r"\(\s*(\d+(?:\.\d+)?)\s*\*\s*0\.1125\s*\*\s*\$?[A-Z]\$?\d+\s*/\s*100\s*\)"
)

# Generic /N divisor finder. Used for batch-size detection in food/potion
# formulas. The first small-N divisor in the formula (left-to-right) is the
# batch divisor; later /100 is the station-fee normaliser, so we cap N at 50.
BATCH_RE = re.compile(r"/\s*(\d+(?:\.\d+)?)")

def parse_station_fee(formula):
    """Returns nutrition value (float) embedded in the station-fee term, or None."""
    if not formula or not isinstance(formula, str):
        return None
    m = STATION_FEE_RE.search(formula)
    return float(m.group(1)) if m else None

def parse_batch_divisor(formula):
    """Returns the batch divisor (e.g. 10 for soups, 5 for potions) or None.
    Scans for the first /N where 1 < N <= 50; this skips the /100 that's
    part of the station-fee normaliser and tolerates a `+ Materials!XY)`
    sitting between (1 - V15) and the batch divisor (the Gathering/Tornado
    potion pattern)."""
    if not formula or not isinstance(formula, str):
        return None
    for m in BATCH_RE.finditer(formula):
        n = float(m.group(1))
        if 1 < n <= 50:
            return n
    return None


# ---------- WALK CRAFT SHEETS ----------
SKIP = {'Intro', 'Instructions', 'Materials', 'Blacksmith', 'Hunter', 'Mage',
        'Refining', 'Food And Potions', 'Tool maker'}

ENCHANT_COLS = {'I': 0, 'K': 1, 'M': 2, 'O': 3, 'Q': 4}

def find_recipe_block_starts(ws):
    """Find rows that start a recipe table. Heuristic: a row that has 'Cost' in column G
    AND has 'Enchantment 0' in column I."""
    starts = []
    for r in range(1, ws.max_row + 1):
        gv = ws.cell(row=r, column=7).value  # G
        iv = ws.cell(row=r, column=9).value  # I
        if isinstance(gv, str) and gv.strip().lower() == 'cost' and isinstance(iv, str) and 'Enchantment 0' in iv:
            # Find the item label: usually column A on the row above 'Cost' or in the next data row's column A
            starts.append(r)
    return starts

def find_all_cost_rows(ws):
    """Return every row that has 'Cost' in column G, regardless of whether
    column I says 'Enchantment 0'. Used as block-boundary markers so that
    a non-standard 'Cost' header (e.g. CapesFurniture's Furniture sub-block
    where column I lists item names like Bed/Chest instead of 'Enchantment 0')
    still terminates the previous recipe block instead of swallowing its
    rows."""
    rows = []
    for r in range(1, ws.max_row + 1):
        gv = ws.cell(row=r, column=7).value
        if isinstance(gv, str) and gv.strip().lower() == 'cost':
            rows.append(r)
    return rows

def extract_sheet(ws):
    """Yield recipes from a craft/refining sheet."""
    sheet_name = ws.title
    starts = find_recipe_block_starts(ws)
    recipes = []
    for header_row in starts:
        # Find the section name (item base name) - search column A from header_row upward
        section = None
        for rr in range(header_row, max(0, header_row - 20), -1):
            v = ws.cell(row=rr, column=1).value
            if isinstance(v, str) and v.strip() and not v.strip().startswith('*'):
                section = v.strip()
                break
        # Also check column G above header for an item-name column (some sheets use G)
        # Iterate rows below header until next header or empty stretch
        end_row = ws.max_row
        for next_start in starts:
            if next_start > header_row:
                end_row = min(end_row, next_start - 1)
                break
        # Look for tier/item rows: column G has a label, columns I/K/M/O/Q have formulas
        for r in range(header_row + 1, end_row + 1):
            tier_label = ws.cell(row=r, column=7).value  # G
            if not isinstance(tier_label, str) or not tier_label.strip():
                continue
            tier_label = tier_label.strip()
            # Check if there's at least one formula
            has_recipe = False
            row_recipes = {}
            for col_letter, ench in ENCHANT_COLS.items():
                cell = ws[f'{col_letter}{r}']
                f = cell.value
                items = parse_formula(f) if isinstance(f, str) and f.startswith('=') else None
                if items:
                    has_recipe = True
                    row_recipes[ench] = items
            if has_recipe:
                # Determine item name. Some sheets put item name in tier_label (e.g. 'Carrot Soup T1'),
                # others put generic 'Tier 4'. If tier_label starts with 'Tier' and section exists, combine.
                if section and tier_label.lower().startswith('tier'):
                    item_name = f'{section} {tier_label}'
                else:
                    item_name = tier_label
                recipes.append({
                    'sheet': sheet_name,
                    'section': section,
                    'item': item_name,
                    'tierLabel': tier_label,
                    'enchantments': row_recipes,
                })
    return recipes


def extract_gathering_gear(ws):
    """Special-case: GatheringGear uses 6 item columns (I/K/M/O/Q/S) per section.
    Header row contains item names; each subsequent even row is a tier with cost formulas.
    Each family (Harvester/Skinner/...) is exposed as its OWN virtual sheet so the
    sidebar lists them as separate menu entries."""
    GEAR_COLS = ['I', 'K', 'M', 'O', 'Q', 'S']
    FAMILY_TO_SHEET = {
        'Harvester':  'GatheringHarvester',
        'Skinner':    'GatheringSkinner',
        'Miner':      'GatheringMiner',
        'Quarrier':   'GatheringQuarrier',
        'Lumberjack': 'GatheringLumberjack',
        'Fisherman':  'GatheringFisherman',
    }
    recipes = []
    # Find header rows (rows where column G has 'Cost')
    headers = []
    for r in range(1, ws.max_row + 1):
        gv = ws.cell(row=r, column=7).value
        if isinstance(gv, str) and gv.strip().lower() == 'cost':
            headers.append(r)

    for hi, hrow in enumerate(headers):
        # Section name from column A on the same row
        section = ws.cell(row=hrow, column=1).value
        section = section.strip() if isinstance(section, str) else f'Section{hi}'
        # Map to virtual sheet name
        virtual_sheet = FAMILY_TO_SHEET.get(section, 'GatheringGear')
        # Item names in I/K/M/O/Q/S of header row
        item_names = {}
        for col in GEAR_COLS:
            v = ws[f'{col}{hrow}'].value
            if isinstance(v, str) and v.strip():
                item_names[col] = v.strip()

        end_row = headers[hi + 1] - 1 if hi + 1 < len(headers) else ws.max_row
        # Walk tier rows
        for r in range(hrow + 1, end_row + 1):
            tier_label = ws.cell(row=r, column=7).value
            if not isinstance(tier_label, str) or not tier_label.strip():
                continue
            tier_label = tier_label.strip()
            for col in GEAR_COLS:
                if col not in item_names:
                    continue
                f = ws[f'{col}{r}'].value
                items = parse_formula(f) if isinstance(f, str) and f.startswith('=') else None
                if items:
                    # Each piece (Cap / Garb / Workboots / Backpack / Tool /
                    # Avalonian Tool) gets its own section so the table renders
                    # one merged image+name cell per piece.
                    recipes.append({
                        'sheet': virtual_sheet,
                        'section': item_names[col],
                        'item': f'{item_names[col]} {tier_label}',
                        'tierLabel': tier_label,
                        'enchantments': {0: items},
                    })
    return recipes


# ---------- ARTIFACT CATALOG PER SHEET ----------
# Most weapon/armor/accessory sheets have up to 6 artifact items in row 5
# (cols I/K/M/O/Q, plus S for CapesFurniture). Tier rows 6/8/10/12/14 = T4-T8.
# CapesFurniture is special: it has 13 cape crests stretching across
# I/K/M/O/Q/S/U/W/Y/AA/AC/AE (6 city + Brecilien + 5 faction). Extending
# the column list catches them all; other sheets are blank past S so the
# extra columns simply produce no entries there.
ARTIFACT_TIER_ROWS = {6: 'T4', 8: 'T5', 10: 'T6', 12: 'T7', 14: 'T8'}
ARTIFACT_COLS = ['I', 'K', 'M', 'O', 'Q', 'S', 'U', 'W', 'Y', 'AA', 'AC', 'AE']

# Per-sheet artifact maps: sheet_name -> { local_coord (e.g. 'I6') -> material_id }
sheet_artifact_maps = {}

def slug(s):
    return re.sub(r'[^A-Z0-9]+', '_', s.upper().strip()).strip('_')

def build_artifact_catalog(wb):
    for sn in wb.sheetnames:
        if sn in SKIP or sn == 'GatheringGear':
            continue
        ws = wb[sn]
        # Read artifact NAMES in row 5
        names = {}  # col_letter -> name
        for col in ARTIFACT_COLS:
            v = ws[f'{col}5'].value
            if isinstance(v, str) and v.strip() and v.strip().upper() != 'ARTIFACTS':
                names[col] = v.strip()
        if not names:
            continue
        local_map = {}
        for col, name in names.items():
            for row, tier in ARTIFACT_TIER_ROWS.items():
                mid = f'ART_{slug(sn)}_{slug(name)}_{tier}'
                local_map[f'{col}{row}'] = mid
                mat_meta[mid] = {
                    'id': mid,
                    'family': f'ARTIFACT_{sn}',
                    'tier': tier,
                    'kind': 'artifact',
                    'name': f'{name} {tier}',
                    'sheet': sn,
                }
        sheet_artifact_maps[sn] = local_map

build_artifact_catalog(wb)


# ---------- FOOD SHEET FISH-PRICE CATALOG ----------
# The Food sheet has its own fish-price grid (not in Materials!).
#   Row 6 headers:  I=Eel  K=Eye  M=Lurcher  O=Crab  Q=Clam  S=Squid  U=Snapper
#   Row 7 = T3   Row 9 = T5   Row 11 = T7  (prices live at I/K/M/O/Q/S/U on those rows)
# Recipes reference these as bare cell refs like "Q7" — no Materials! prefix.
FISH_COLS = {'I':'Eel','K':'Eye','M':'Lurcher','O':'Crab','Q':'Clam','S':'Squid','U':'Snapper'}
FISH_TIER_ROWS = {7: 'T3', 9: 'T5', 11: 'T7'}
ws_food = wb['Food'] if 'Food' in wb.sheetnames else None
if ws_food is not None:
    food_local_map = sheet_artifact_maps.setdefault('Food', {})
    for col, name in FISH_COLS.items():
        for row, tier in FISH_TIER_ROWS.items():
            mid = f'FISH_{slug(name)}_{tier}'
            food_local_map[f'{col}{row}'] = mid
            mat_meta[mid] = {
                'id': mid,
                'family': 'FISH',
                'subFamily': name.upper(),
                'tier': tier,
                'kind': 'fish',
                'name': f'{name} {tier}',
            }


# ---------- UPDATE FORMULA PARSER FOR LOCAL REFS ----------
# Match $-anchored OR bare cell refs that aren't prefixed with `Materials!`.
# Examples: $I$6 (artifact, Sword sheet), Q7 (fish, Food sheet),
# AE6 (Demon Crest artifact on CapesFurniture — uses a double-letter
# column past Z). The `[A-Z]{1,2}` class catches both single and double
# letter columns; `\b` boundary stops a stray `E6` mid-formula from being
# misparsed as the cell `E6`.
LOCAL_REF = re.compile(r"\$?\b([A-Z]{1,2})\$?(\d+)")

def parse_formula_with_artifacts(formula, sheet_name):
    """Wraps parse_formula; additionally detects local artifact refs."""
    items = parse_formula(formula) or []
    if not formula or not isinstance(formula, str):
        return items if items else None
    f = formula.strip()
    if not f.startswith('='):
        return items if items else None

    sheet_map = sheet_artifact_maps.get(sheet_name, {})
    if not sheet_map:
        return items if items else None
    for m in LOCAL_REF.finditer(f):
        coord = f'{m.group(1)}{m.group(2)}'
        # Skip refs prefixed with `Materials!` (already covered above)
        start = m.start()
        if start >= len('Materials!'):
            preceding = f[max(0, start - len('Materials!')):start]
            if preceding.endswith('Materials!'):
                continue
        mid = sheet_map.get(coord)
        if not mid:
            continue
        # Avoid duplicates
        if any(it['mat'] == mid for it in items):
            continue
        # Artifacts AND fish are added outside the (1-returnFactor) bracket
        # for artifacts, but fish ARE inside the bracket in the food formulas
        # (e.g. `(2 * Materials!$D$151 + Q7) * (1 - V15)`). Flag accordingly.
        if mid.startswith('FISH_'):
            items.append({'mat': mid, 'qty': 1})
        else:
            items.append({'mat': mid, 'qty': 1, 'noReturnDiscount': True})
    return items if items else None


def extract_sheet_v2(ws):
    """Like extract_sheet, but uses parse_formula_with_artifacts and also
    captures the station-fee nutrition cost (Food/Potions only).
    Skips sections whose name ends with '*' — those are the refining sheets'
    duplicate "Refine*" tables that compute the same recipes using the user's
    own previous-tier output instead of market prices. The web UI handles
    chain-refining dynamically, so the duplicate data is unnecessary."""
    sheet_name = ws.title
    starts = find_recipe_block_starts(ws)
    all_costs = find_all_cost_rows(ws)
    recipes = []
    for header_row in starts:
        # Find the section label in column A. Priorities (closest wins):
        #   (1) header_row itself
        #   (2) the row immediately below the header — some sections in
        #       the Nendys sheet (e.g. Potions / Poison) put the section
        #       name on the first data row, NOT above the Cost header,
        #       and the old upward-only search would inherit the previous
        #       section's name ('Resistance') for Poison.
        #   (3) walk upward (Soups / Salads style, where the label sits
        #       a couple of rows ABOVE the Cost header)
        section = None
        def _cell_label(r):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip() and not v.strip().startswith('*'):
                return v.strip()
            return None
        section = _cell_label(header_row) or _cell_label(header_row + 1)
        if not section:
            for rr in range(header_row - 1, max(0, header_row - 20), -1):
                lbl = _cell_label(rr)
                if lbl:
                    section = lbl
                    break
        # Skip the asterisk variants that refining sheets use to express
        # "compute as if you crafted the previous tiers yourself". The
        # calculator handles chain-refining at runtime.
        if section and section.endswith('*'):
            continue
        # End the block at the next 'Cost' row of ANY layout — not just
        # standard Enchantment-0 blocks. This catches CapesFurniture, where
        # the Furniture sub-block (Bed / Chest / Table) uses 'Cost' in G
        # but lists item names in I, and would otherwise let the previous
        # cape section absorb the Bed rows.
        end_row = ws.max_row
        for next_cost in all_costs:
            if next_cost > header_row:
                end_row = min(end_row, next_cost - 1)
                break
        for next_start in starts:
            if next_start > header_row:
                end_row = min(end_row, next_start - 1)
                break
        for r in range(header_row + 1, end_row + 1):
            tier_label = ws.cell(row=r, column=7).value
            if not isinstance(tier_label, str) or not tier_label.strip():
                continue
            tier_label = tier_label.strip()
            has_recipe = False
            row_recipes = {}
            row_iv = {}     # enchant -> Item Value (extracted from formula)
            row_batch = {}  # enchant -> batch divisor
            for col_letter, ench in ENCHANT_COLS.items():
                cell = ws[f'{col_letter}{r}']
                f = cell.value
                items = (parse_formula_with_artifacts(f, sheet_name)
                         if isinstance(f, str) and f.startswith('=') else None)
                if items:
                    has_recipe = True
                    row_recipes[ench] = items
                    iv = parse_station_fee(f)
                    if iv is not None:
                        row_iv[ench] = iv
                    div = parse_batch_divisor(f)
                    if div is not None:
                        row_batch[ench] = div
            if has_recipe:
                if section and tier_label.lower().startswith('tier'):
                    item_name = f'{section} {tier_label}'
                else:
                    item_name = tier_label
                rec = {
                    'sheet': sheet_name,
                    'section': section,
                    'item': item_name,
                    'tierLabel': tier_label,
                    'enchantments': row_recipes,
                }
                if row_iv:
                    rec['iv'] = row_iv
                if row_batch:
                    rec['batch'] = row_batch
                recipes.append(rec)
    return recipes


def extract_bags_misc_block(ws):
    """The BagsSatchelsTracking sheet has a 'Miscellaneous' block whose layout
    matches GatheringGear (5 items in cols I/K/M/O/Q with tier rows below)
    instead of the usual Enchantment 0-4 pattern. extract_sheet_v2 misses it
    because it looks for 'Enchantment 0' in column I.

    Header row 60: I=Tracking Kit, K=Siege Banner, M=Siege Hammer,
                   O=Avalonian Hammer, Q=Repair Kits.
    Tier rows: 61/63/65/67/69/71/73 = T2..T8.
    """
    GEAR_COLS = ['I', 'K', 'M', 'O', 'Q']
    # Find the header row by content.
    header_row = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        i = ws.cell(row=r, column=9).value
        if isinstance(a, str) and 'Miscellaneous' in a and isinstance(i, str) and i.strip():
            header_row = r; break
    if not header_row:
        return []

    item_names = {}
    for col in GEAR_COLS:
        v = ws[f'{col}{header_row}'].value
        if isinstance(v, str) and v.strip():
            item_names[col] = v.strip()

    recipes = []
    for r in range(header_row + 1, ws.max_row + 1):
        tier_label = ws.cell(row=r, column=7).value
        if not isinstance(tier_label, str) or not tier_label.strip():
            continue
        tier_label = tier_label.strip()
        if not re.match(r'^Tier\s*\d+', tier_label, re.IGNORECASE):
            break  # left the misc block
        for col in GEAR_COLS:
            if col not in item_names:
                continue
            f = ws[f'{col}{r}'].value
            items = parse_formula(f) if isinstance(f, str) and f.startswith('=') else None
            if items:
                recipes.append({
                    'sheet': 'BagsSatchelsTracking',
                    'section': item_names[col],
                    'item': f'{item_names[col]} {tier_label}',
                    'tierLabel': tier_label,
                    'enchantments': {0: items},
                })
    return recipes


all_recipes = []
for sn in wb.sheetnames:
    if sn in SKIP:
        continue
    ws = wb[sn]
    if sn == 'GatheringGear':
        rs = extract_gathering_gear(ws)
    else:
        rs = extract_sheet_v2(ws)
    # BagsSatchelsTracking has an extra block below the standard Bags / Satchels
    # sections that the v2 extractor doesn't pick up — pull it in manually.
    if sn == 'BagsSatchelsTracking':
        rs = rs + extract_bags_misc_block(ws)
    all_recipes.extend(rs)
    print(f'{sn}: {len(rs)} recipes')

# ---------- RECIPE-FAMILY OVERRIDES (Nendys spreadsheet bugs) ----------
# A handful of sections in the source spreadsheet reference the wrong refined
# material family. We substitute on the way out so the calculator matches the
# official Albion craftingrequirements (verified against ao-data/ao-bin-dumps
# items.json). Quantities and tiers are kept; only the FAMILY changes — e.g.
# all Quarterstaff recipes have their PLANKS_T4.0 swapped to STEEL_T4.0.
#
# Format: { (sheet, section_or_None_for_all): {old_family: new_family} }
RECIPE_FAMILY_FIXES = {
    # Spreadsheet had PLANKS where Albion uses STEEL (Warrior weapon).
    ('Quarterstaffs', None):                       {'PLANKS': 'STEEL'},
    # Avalonian leather armor extracted with STEEL — needs LEATHER.
    ('LeatherJackets', 'Jacket of Tenacity'):      {'STEEL':  'LEATHER'},
    # Avalonian cloth armor extracted with STEEL — needs CLOTH.
    ('ClothRobes',     'Robe of Purity'):          {'STEEL':  'CLOTH'},
    # Assassin leather pieces had their material column set to STEEL ($I)
    # in the Nendys spreadsheet instead of LEATHER ($N). Jacket is fine.
    ('LeatherHoods',   'Assassin Hood'):           {'STEEL':  'LEATHER'},
    ('LeatherShoes',   'Assassin Shoes'):          {'STEEL':  'LEATHER'},
    # Fishing rod extracted with STEEL — needs CLOTH (per official recipe).
    ('GatheringFisherman', 'Rod'):                 {'STEEL':  'CLOTH'},
}

def apply_family_fix(rec):
    """Look up override for this recipe's (sheet, section) and rewrite material
    ids in every enchantment row in place."""
    rules = RECIPE_FAMILY_FIXES.get((rec['sheet'], rec.get('section')))
    if not rules:
        rules = RECIPE_FAMILY_FIXES.get((rec['sheet'], None))
    if not rules:
        return
    for ench, items in rec['enchantments'].items():
        for it in items:
            mid = it['mat']
            for old_fam, new_fam in rules.items():
                if mid.startswith(old_fam + '_'):
                    new_mid = new_fam + mid[len(old_fam):]
                    if new_mid in mat_meta:
                        it['mat'] = new_mid
                    break

for rec in all_recipes:
    apply_family_fix(rec)


# ---------- POST-PROCESSING FIXES (qty / ingredient mismatches) ----------
# These can't be fixed by a simple family swap — they involve quantity
# overrides or adding/removing materials entirely. Each function mutates
# a single enchantment's items list.

def _fix_infinity_blade(items):
    """Spreadsheet hard-codes the 2H ratio (20 STEEL + 12 LEATHER). The actual
    Albion item T4_MAIN_SWORD_CRYSTAL is a 1-handed sword with 16 STEEL +
    8 LEATHER. Override quantities; family stays the same."""
    for it in items:
        if it['mat'].startswith('STEEL_'):   it['qty'] = 16
        elif it['mat'].startswith('LEATHER_'): it['qty'] = 8

def _fix_gathering_backpack(items):
    """Spreadsheet has only 4 LEATHER; Albion needs 4 CLOTH + 4 LEATHER."""
    leather = next((it for it in items if it['mat'].startswith('LEATHER_')), None)
    if not leather: return
    cloth_id = 'CLOTH' + leather['mat'][len('LEATHER'):]
    if cloth_id in mat_meta and not any(it['mat'] == cloth_id for it in items):
        items.append({'mat': cloth_id, 'qty': leather['qty']})

def _fix_avalonian_hammer(items):
    """Spreadsheet: 8 BLOCKS + 8 PLANKS + 20 avalon-energy.
    Official:      6 PLANKS + 2 STEEL + 20 avalon-token.
    Drop BLOCKS, add STEEL at qty 2, set PLANKS to 6."""
    items[:] = [it for it in items if not it['mat'].startswith('BLOCKS_')]
    plank = next((it for it in items if it['mat'].startswith('PLANKS_')), None)
    if not plank: return
    plank['qty'] = 6
    steel_id = 'STEEL' + plank['mat'][len('PLANKS'):]
    if steel_id in mat_meta and not any(it['mat'] == steel_id for it in items):
        items.append({'mat': steel_id, 'qty': 2})

def _swap_axes_crystal_avalon(items):
    """The Nendys spreadsheet's Axes sheet has Crystal Reaper and Realmbreaker
    pointing at each other's artifact: Crystal Reaper formula references the
    Avalonian Battle Memoir column, and Realmbreaker references Edged Crystal.
    Swap the two artifact ids in place so each section uses the right
    artifact for its in-game item (T4_2H_SCYTHE_CRYSTAL = Edged Crystal;
    T4_2H_AXE_AVALON = Avalonian Battle Memoir)."""
    SWAP = {
        'ART_AXES_AVALONIAN_BATTLE_MEMOIR_': 'ART_AXES_EDGED_CRYSTAL_',
        'ART_AXES_EDGED_CRYSTAL_':           'ART_AXES_AVALONIAN_BATTLE_MEMOIR_',
    }
    for it in items:
        if not it.get('noReturnDiscount'):
            continue
        for old_prefix, new_prefix in SWAP.items():
            if it['mat'].startswith(old_prefix):
                it['mat'] = new_prefix + it['mat'][len(old_prefix):]
                break

POST_FIXES = {
    ('Swords', 'Infinity Blade'):                 _fix_infinity_blade,
    ('GatheringHarvester',  'Harvester Backpack'):  _fix_gathering_backpack,
    ('GatheringSkinner',    'Skinner Backpack'):    _fix_gathering_backpack,
    ('GatheringMiner',      'Miner Backpack'):      _fix_gathering_backpack,
    ('GatheringQuarrier',   'Quarrier Backpack'):   _fix_gathering_backpack,
    ('GatheringLumberjack', 'Lumberjack Backpack'): _fix_gathering_backpack,
    ('GatheringFisherman',  'Fisherman Backpack'):  _fix_gathering_backpack,
    ('GatheringQuarrier',     'Avalonian Hammer'):  _fix_avalonian_hammer,
    ('BagsSatchelsTracking',  'Avalonian Hammer'):  _fix_avalonian_hammer,
    ('Axes', 'Crystal Reaper'):  _swap_axes_crystal_avalon,
    ('Axes', 'Realmbreaker'):    _swap_axes_crystal_avalon,
}

for rec in all_recipes:
    fix = POST_FIXES.get((rec['sheet'], rec.get('section')))
    if not fix: continue
    for ench, items in rec['enchantments'].items():
        fix(items)


# Stone Refining keeps its full enchant-column set. The Nendys formulas
# are correct as written: refining with enchant-N raw rocks produces base
# (non-enchanted) stoneblocks, and the columns let the user compare the
# silver-per-craft cost of using base / uncommon / rare / exceptional
# rocks side-by-side. The Stone Refining page renames the column headers
# to reflect that ("Base / Uncommon / Rare / Exceptional" raws) instead
# of "Enchantment 0/1/2/3".


# ---------- MERGE FURNITURE RECIPES ----------
# Furniture isn't in the Nendys spreadsheet — recipes are sourced separately
# from ao-data/ao-bin-dumps via extract_furniture.py and live in
# albion/furniture.json. We splice them into all_recipes here so the rest
# of the pipeline (IV stamping, JSON output) treats them uniformly.
import os as _os
furniture_path = _os.path.join(_os.path.dirname(OUT), 'furniture.json')
if _os.path.exists(furniture_path):
    with open(furniture_path, encoding='utf-8') as _f:
        _fdata = json.load(_f)
    for _rec in _fdata.get('recipes', []):
        # JSON keys come out as strings; normalise enchantments to int keys
        # so the rest of the pipeline matches existing recipes.
        _ench = {}
        for _k, _v in _rec.get('enchantments', {}).items():
            try: _ench[int(_k)] = _v
            except: _ench[_k] = _v
        _rec['enchantments'] = _ench
        all_recipes.append(_rec)
    print(f'Furniture: {len(_fdata.get("recipes", []))} recipes merged in')


# ---------- STAMP IV ON EVERY MATERIAL ----------
for mid, meta in mat_meta.items():
    tier_for_iv = meta.get('tier')
    # Food / potion materials don't carry the tier in the `tier` field
    # (they use 'T-' or '-'); pull it from the name when possible.
    if not tier_for_iv or tier_for_iv == '-':
        tier_for_iv = parse_tier_from_name(meta.get('name'))
    meta['iv'] = derive_iv(tier_for_iv) if tier_for_iv else 0

# ---------- FALLBACK IV ON EACH RECIPE ----------
# Spreadsheet IV (parsed from the +(IV * 0.1125 * Fee/100) station-fee term)
# only exists on Food / Potion T3+ recipes. For every OTHER recipe — weapons,
# armor, gathering gear, T1/T2 food and potions — synthesise an IV by summing
# qty * material.iv across the recipe's non-heart, non-artifact ingredients,
# matching Albion's `IV = NumItems * Base` rule for non-artifact crafts.
#
# Refining recipes get a different rule: the OUTPUT item's IV is just the
# refined-material IV at that tier+ench, not the sum of input IVs.
REFINING_TO_FAMILY = {
    'PlankRefining': 'PLANKS', 'SteelRefining': 'STEEL',
    'LeatherRefining': 'LEATHER', 'ClothRefining': 'CLOTH',
    'StoneRefining': 'BLOCKS',
}

def refining_output_iv(rec, ench):
    """For refining recipes, the IV is the output material's IV at the
    appropriate tier+ench. tier_label looks like 'Tier 4' or 'Tier 7'."""
    family = REFINING_TO_FAMILY.get(rec['sheet'])
    if not family:
        return 0
    import re as _re
    m = _re.match(r'tier\s*(\d+)', (rec['tierLabel'] or '').strip(), _re.IGNORECASE)
    if not m:
        return 0
    base_tier = int(m.group(1))
    e_int = int(ench) if isinstance(ench, str) and ench.isdigit() else (ench if isinstance(ench, int) else 0)
    return 16 * (2 ** (base_tier - 4)) * (2 ** e_int)


# ---------- HAND-AUTHORED ADDITIONS (post-spreadsheet) ----------
# Items added to Albion after the source spreadsheet was last updated.
# These are inlined here rather than carried in data.json so re-extracting
# from a fresh xlsx doesn't drop them.
#
# IMPORTANT: this block runs BEFORE the IV-computation loop below so that
# recipes without an explicit iv get auto-computed from their ingredient
# IVs (sum of qty * mat.iv). Recipes that DO set iv (like Smuggler Cape)
# are preserved by the loop's "spreadsheet value wins" rule.

# Smuggler's Cape — Brecilien-themed faction cape using Shadowheart +
# its own crest artifact. Same shape as the other artifact capes: per
# tier 4×Cloth + 4×Leather + 1×Shadowheart + 1×Smuggler Crest, with
# enchantment levels 0-4 scaling cloth/leather tier only.
SMUGGLER_TIERS = ['T4', 'T5', 'T6', 'T7', 'T8']
for t in SMUGGLER_TIERS:
    mid = f'ART_CAPESFURNITURE_SMUGGLER_CREST_{t}'
    mat_meta[mid] = {
        'id': mid,
        'family': 'ARTIFACT_CapesFurniture',
        'tier': t,
        'kind': 'artifact',
        'name': f'Smuggler Crest {t}',
        'sheet': 'CapesFurniture',
    }

def _smuggler_recipe(tier_num):
    crest = f'ART_CAPESFURNITURE_SMUGGLER_CREST_T{tier_num}'
    ench = {}
    for e in range(5):
        ench[str(e)] = [
            {'mat': f'CLOTH_T{tier_num}.{e}',   'qty': 4.0},
            {'mat': f'LEATHER_T{tier_num}.{e}', 'qty': 4.0},
            {'mat': 'HEART_SHADOWHEART', 'qty': 1, 'noReturnDiscount': True},
            {'mat': crest, 'qty': 1, 'noReturnDiscount': True},
        ]
    return {
        'sheet': 'CapesFurniture',
        'section': 'Smuggler Cape',
        'item': f'Smuggler Cape Tier {tier_num}',
        'tierLabel': f'Tier {tier_num}',
        'enchantments': ench,
    }
for t in [4, 5, 6, 7, 8]:
    all_recipes.append(_smuggler_recipe(t))


# Avalonian Cape — post-spreadsheet artifact cape that uses Avalonian
# Energy instead of a heart. Recipe shape per Albion: 4×Cloth + 4×Leather
# + N×Avalonian Energy + 1×Avalonian Cape artifact. Energy quantities
# per tier (user-confirmed): T4=15, T5=15, T6=45, T7=75, T8=150.
AVALON_TIERS = ['T4', 'T5', 'T6', 'T7', 'T8']
for t in AVALON_TIERS:
    mid = f'ART_CAPESFURNITURE_AVALONIAN_CAPE_{t}'
    mat_meta[mid] = {
        'id': mid,
        'family': 'ARTIFACT_CapesFurniture',
        'tier': t,
        'kind': 'artifact',
        'name': f'Avalonian Cape {t}',
        'sheet': 'CapesFurniture',
    }
_AVALON_ENERGY_PER_TIER = {4: 15, 5: 15, 6: 45, 7: 75, 8: 150}
def _avalon_cape_recipe(tier_num):
    artifact = f'ART_CAPESFURNITURE_AVALONIAN_CAPE_T{tier_num}'
    energy_qty = _AVALON_ENERGY_PER_TIER[tier_num]
    ench = {}
    for e in range(5):
        ench[str(e)] = [
            {'mat': f'CLOTH_T{tier_num}.{e}',   'qty': 4.0},
            {'mat': f'LEATHER_T{tier_num}.{e}', 'qty': 4.0},
            {'mat': 'MISC_AVALONIAN_ENERGY', 'qty': float(energy_qty), 'noReturnDiscount': True},
            {'mat': artifact, 'qty': 1, 'noReturnDiscount': True},
        ]
    return {
        'sheet': 'CapesFurniture',
        'section': 'Avalonian Cape',
        'item': f'Avalonian Cape Tier {tier_num}',
        'tierLabel': f'Tier {tier_num}',
        'enchantments': ench,
    }
for t in [4, 5, 6, 7, 8]:
    all_recipes.append(_avalon_cape_recipe(t))


# Roasts (Food sheet) — meat-heavy meal added to Albion after the
# spreadsheet's last update. Only T3 / T5 / T7 exist in-game (no
# even-tier roasts). User-confirmed recipes:
#
#   T3 Roast Chicken: 8  Raw Chicken + 4  T2 Beans   + 4  T4 Goat's Milk
#   T5 Roast Goose:   24 Raw Goose   + 12 T5 Cabbage + 12 T6 Sheep's Milk
#   T7 Roast Pork:    72 Raw Pork    + 36 T7 Corn    + 36 T8 Cow's Milk
#
# Enchanted variants add fish sauce on top of the base recipe, matching
# the qty pattern of other meat meals at the same tier (Pork Omelette,
# Beef Stew, etc.): T3=10, T5=30, T7=90 per enchant level.
#
# batch=10 (matches other food meals); IV is auto-computed by the loop
# below from each ingredient's IV.
_ROASTS = [
    # (tier, item-name,     meat-id,                   veg-id,                         milk-id,                    meat-qty, veg-qty, milk-qty, sauce-qty)
    (3, 'Roast Chicken', 'FP_TIER_3___RAW_CHICKEN', 'FP_TIER_2___BEANS',          'FP_TIER_4___GOATS_MILK',    8,  4,  4, 10),
    (5, 'Roast Goose',   'FP_TIER_5___RAW_GOOSE',   'FP_TIER_5___CABBAGE',        'FP_TIER_6___SHEEPS_MILK',  24, 12, 12, 30),
    (7, 'Roast Pork',    'FP_TIER_7___RAW_PORK',    'FP_TIER_7___BUNDLE_OF_CORN', 'FP_TIER_8___COWS_MILK',    72, 36, 36, 90),
]
_FISH_SAUCE_BY_ENCH = {
    1: 'FP_BASIC_FISH_SAUCE',
    2: 'FP_FANCY_FISH_SAUCE',
    3: 'FP_SPECIAL_FISH_SAUCE',
}
for tier, item_name, meat, veg, milk, mq, vq, kq, sq in _ROASTS:
    base = [
        {'mat': meat, 'qty': float(mq)},
        {'mat': veg,  'qty': float(vq)},
        {'mat': milk, 'qty': float(kq)},
    ]
    ench = {'0': list(base)}
    batch = {'0': 10.0}
    for e, sauce_id in _FISH_SAUCE_BY_ENCH.items():
        ench[str(e)] = list(base) + [{'mat': sauce_id, 'qty': float(sq)}]
        batch[str(e)] = 10.0
    all_recipes.append({
        'sheet': 'Food',
        'section': 'Roasts',
        'item': f'{item_name} T{tier}',
        'tierLabel': f'{item_name} T{tier}',
        'enchantments': ench,
        'batch': batch,
    })

# Fish-roast variants (same Roasts section). Rare-snapper-themed dishes:
#   T3 Roasted Whitefrog Snapper:  1 Snapper T3 + 1 Comfrey         + 1 T4 Goat's Milk
#   T5 Roasted Clearhaze Snapper:  1 Snapper T5 + 2 Cabbage + 2 Dragon Teasel + 2 T6 Sheep's Milk
#   T7 Roasted Puremist Snapper:   1 Snapper T7 + 6 Corn + 6 Firetouched Mullein + 6 T8 Cow's Milk
# Enchant fish-sauce qty follows the seafood-meal pattern (Eel Stew /
# Squid Salad / Lurcher Sandwich): 3 / 9 / 27 by tier — smaller than
# meat-meal sauce qty because the base recipe quantities are smaller.
_FISH_ROASTS = [
    # (tier, item-name,                    fish-id,            extra-ingredients [list of (mat, qty)], sauce-qty)
    (3, 'Roasted Whitefrog Snapper', 'FISH_SNAPPER_T3', [
        ('FP_TIER_3___BRIGHTLEAF_COMFREY', 1),
        ('FP_TIER_4___GOATS_MILK',          1),
    ], 3),
    (5, 'Roasted Clearhaze Snapper', 'FISH_SNAPPER_T5', [
        ('FP_TIER_5___CABBAGE',       2),
        ('FP_TIER_5___DRAGON_TEASEL', 2),
        ('FP_TIER_6___SHEEPS_MILK',   2),
    ], 9),
    (7, 'Roasted Puremist Snapper',  'FISH_SNAPPER_T7', [
        ('FP_TIER_7___BUNDLE_OF_CORN',         6),
        ('FP_TIER_7___FIRETOUCHED_MULLEIN',    6),
        ('FP_TIER_8___COWS_MILK',              6),
    ], 27),
]
for tier, item_name, fish_id, extras, sq in _FISH_ROASTS:
    base = [{'mat': fish_id, 'qty': 1.0}] + [
        {'mat': m, 'qty': float(q)} for m, q in extras
    ]
    ench = {'0': list(base)}
    for e, sauce_id in _FISH_SAUCE_BY_ENCH.items():
        ench[str(e)] = list(base) + [{'mat': sauce_id, 'qty': float(sq)}]
    # Snapper roasts output 1 unit per craft (like every other fish meal:
    # Eel Stew, Squid Salad, Clam Soup, etc.) — do NOT set batch=10 here.
    # Missing batch defaults to 1 in the app's cost formula.
    all_recipes.append({
        'sheet': 'Food',
        'section': 'Roasts',
        'item': f'{item_name} T{tier}',
        'tierLabel': f'{item_name} T{tier}',
        'enchantments': ench,
    })


for rec in all_recipes:
    iv_map = rec.get('iv', {})
    is_refining = rec['sheet'] in REFINING_TO_FAMILY
    for ench, items in rec['enchantments'].items():
        if ench in iv_map and iv_map[ench]:
            continue  # spreadsheet value wins
        if is_refining:
            iv_map[ench] = refining_output_iv(rec, ench)
            continue
        total = 0
        for it in items:
            if it.get('heartGated'):    continue
            if it.get('noReturnDiscount'):  # artifacts handled by IV multiplier IRL
                continue
            mat = mat_meta.get(it['mat'])
            if not mat:
                continue
            total += (it.get('qty', 0) or 0) * (mat.get('iv', 0) or 0)
        if total > 0:
            iv_map[ench] = total
    if iv_map:
        rec['iv'] = iv_map


# Expand abbreviated names in the Food sheet — the spreadsheet uses
# "Ava" as shorthand for Avalonian variants (Ava Stew, Ava Omelette,
# Ava Goat Sandwich, etc.). Spell it out so the UI label is clearer.
for r in all_recipes:
    if r.get('sheet') != 'Food':
        continue
    for field in ('item', 'tierLabel'):
        v = r.get(field)
        if isinstance(v, str) and v.startswith('Ava '):
            r[field] = 'Avalonian ' + v[4:]

# Correct spreadsheet-side artifact name typos. Only the displayed `name`
# field is rewritten — the underlying material ID is kept so any prices
# the user has already saved stay attached to the right artifact.
_ARTIFACT_NAME_FIXES = {
    'Infernal Cloth Binding': 'Infernal Cloth Visor',
}
for m in mat_meta.values():
    if m.get('kind') != 'artifact':
        continue
    name = m.get('name', '')
    for wrong, right in _ARTIFACT_NAME_FIXES.items():
        if name.startswith(wrong + ' '):  # match "Wrong Name T4" etc.
            m['name'] = right + name[len(wrong):]
            break

# Safety net: certain materials are NEVER refunded by Albion's return
# rate (artifacts, hearts, special drops like Shadow Claws, energies,
# Royal Sigils). The formula parser flags most of these automatically
# from their bracket position, but spreadsheet inconsistencies can
# slip through — e.g. Avalonian Hammer T8 has Royal Sigils inside the
# discount bracket. Force the flag on by material-id prefix so no
# combination of bracket placement can ever miss them.
def _is_never_refundable(mat_id):
    if not isinstance(mat_id, str):
        return False
    if mat_id.startswith(('ART_', 'HEART_', 'MISC_')):
        return True
    # Rare drop catalysts — not crafted, can't be returned. Arcane
    # Extracts and Rare Animal Remains are crafted ingredients and DO
    # get return-rate discount, so they're NOT in this list.
    NEVER_REFUND_SUBSTR = (
        'SHADOW_CLAW', 'WEREWOLF_FANG', 'SPIRIT_PAW',
        'RUNESTONE_TOOTH', 'SYLVIAN_ROOT', 'DAWNFEATHER',
        'IMPS_HORN', 'IMP_HORN', 'IMPS_HORNS',
    )
    return any(s in mat_id for s in NEVER_REFUND_SUBSTR)

# Force-clear the noReturnDiscount flag on Arcane Extracts everywhere —
# the spreadsheet's formula structure left some instances flagged. In
# Albion these crafted catalysts ARE refunded by return rate, same as
# herbs and other base ingredients.
_FORCE_REFUNDABLE = ('ARCANE_EXTRACT', 'RARE_ANIMAL_REMANINGS')
for rec in all_recipes:
    for items in rec.get('enchantments', {}).values():
        for it in items:
            mat = it.get('mat', '')
            if any(s in mat for s in _FORCE_REFUNDABLE):
                if it.get('noReturnDiscount'):
                    del it['noReturnDiscount']

for rec in all_recipes:
    for items in rec.get('enchantments', {}).values():
        for it in items:
            if _is_never_refundable(it.get('mat')) and not it.get('noReturnDiscount'):
                it['noReturnDiscount'] = True


# Pristine (.4) heart bug — the spreadsheet uses HEART_VINEHEART for
# Plank and Leather T_.4 columns, but in-game each refining type uses
# its NATIVE heart at every enchant level: Plank=Treeheart,
# Leather=Beastheart (Steel keeps Mountainheart correctly; Cloth's
# native IS Vineheart so it's consistent; Stone has no heart on .1-.3).
# Rewrite the e4 heart back to the sheet's native heart.
_NATIVE_HEART_BY_SHEET = {
    'PlankRefining':   'HEART_TREEHEART',
    'LeatherRefining': 'HEART_BEASTHEART',
}
for rec in all_recipes:
    native = _NATIVE_HEART_BY_SHEET.get(rec.get('sheet'))
    if not native:
        continue
    ench = rec.get('enchantments', {})
    # enchantment keys may be either int (during extraction) or str
    # (after JSON round-trip); check both.
    items = ench.get(4) if 4 in ench else ench.get('4', [])
    for it in items or []:
        if it.get('mat') == 'HEART_VINEHEART':
            it['mat'] = native


# Tier-7 raw-quantity bug — the spreadsheet hardcodes 4 raws at T7 for
# Plank / Steel / Leather / Cloth refining ("MAX(4 - IF(hearts,1,0), 0)
# * raw"), but Albion's actual T7 refining needs 5 raws (reduced to 4
# with hearts). The same spreadsheet correctly has Stone Refining T7 =
# 5 raws, so this is just an oversight on the other four sheets.
# Rewrite raw qty 4 → 5 on T7 rows. Heart logic (heartReducesQty) still
# applies on top, so Hearts ON drops it back to 4.
_T7_RAW_FAMILY = {
    'PlankRefining':   'LOGS',
    'SteelRefining':   'ORE',
    'LeatherRefining': 'HIDE',
    'ClothRefining':   'FIBER',
}
for rec in all_recipes:
    fam = _T7_RAW_FAMILY.get(rec.get('sheet'))
    if not fam:
        continue
    if 'Tier 7' not in rec.get('item', ''):
        continue
    for items in rec.get('enchantments', {}).values():
        for it in items:
            if it.get('mat', '').startswith(fam + '_T7') and it.get('qty') == 4:
                it['qty'] = 5.0


# Steel Refining prev-tier ingredient bug — the spreadsheet author copied
# the T(N-1).0 base-steel cell reference across every enchant column and
# only updated the T(N-1).4 reference in the last column. Result: T5-T8
# refining at e1/e2/e3 incorrectly demands base prev-tier steel instead
# of the matching enchanted bar. Every other refining sheet (Plank,
# Leather, Cloth) chains matching enchants correctly; Steel is the
# outlier. Rewrite STEEL_T{N-1}.0 → STEEL_T{N-1}.{E} for affected cells
# so the recipe matches Albion's actual game requirement.
import re as _re_steel
for rec in all_recipes:
    if rec.get('sheet') != 'SteelRefining':
        continue
    m = _re_steel.search(r'Tier\s*(\d+)', rec.get('item', ''))
    if not m:
        continue
    tier = int(m.group(1))
    if tier < 5:  # T3/T4 have no enchant variants in the prev-tier slot
        continue
    prev = tier - 1
    for ench_str, items in rec.get('enchantments', {}).items():
        try:
            e = int(ench_str)
        except (TypeError, ValueError):
            continue
        if e == 0 or e == 4:  # e0 already T(prev).0, e4 already T(prev).4
            continue
        correct = f'STEEL_T{prev}.{e}'
        wrong   = f'STEEL_T{prev}.0'
        for it in items:
            if it.get('mat') == wrong:
                it['mat'] = correct

# Group recipes by sheet
by_sheet = {}
for r in all_recipes:
    by_sheet.setdefault(r['sheet'], []).append(r)

# Write JSON
import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)

out = {
    'materials': list(mat_meta.values()),
    'recipes': all_recipes,
    'sheets': sorted(by_sheet.keys()),
}
with open(OUT, 'w', encoding='utf-8') as fh:
    json.dump(out, fh, indent=2, ensure_ascii=False)

print(f'\nMaterials: {len(mat_meta)}')
print(f'Recipes:   {len(all_recipes)}')
print(f'Sheets:    {len(by_sheet)}')
print(f'Saved to:  {OUT}')
